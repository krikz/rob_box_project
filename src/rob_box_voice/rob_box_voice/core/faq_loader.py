"""Utilities for loading event FAQ data from structured files."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List

QUESTION_KEYS = {"question", "вопрос"}
ANSWER_KEYS = {"answer", "ответ"}
CATEGORY_KEYS = {"question type", "question_type", "category", "тип вопроса"}
SOURCE_KEYS = {"source", "источник"}


def load_faq_items(path: str | Path) -> List[Dict[str, str]]:
    """Load FAQ items from `.xlsx`, `.csv`, or `.json` files.

    Args:
        path: Input file path.

    Returns:
        Normalized FAQ rows with `question`, `answer`, `category`, `source` keys.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the file format is unsupported or required columns are missing.
        ImportError: If `.xlsx` loading is requested without `openpyxl` installed.
    """
    source_path = Path(path)
    if not source_path.exists():
        raise FileNotFoundError(f"FAQ file not found: {source_path}")

    suffix = source_path.suffix.lower()
    if suffix == ".csv":
        rows = _load_csv_rows(source_path)
    elif suffix == ".json":
        rows = _load_json_rows(source_path)
    elif suffix == ".xlsx":
        rows = _load_xlsx_rows(source_path)
    else:
        raise ValueError(f"Unsupported FAQ file format: {source_path.suffix}")

    return _normalize_rows(rows, default_source=source_path.name)


def _load_csv_rows(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader]


def _load_json_rows(path: Path) -> List[Dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return [dict(item) for item in payload]
    if isinstance(payload, dict):
        if isinstance(payload.get("items"), list):
            return [dict(item) for item in payload["items"]]
        if isinstance(payload.get("faq"), list):
            return [dict(item) for item in payload["faq"]]
    raise ValueError("JSON FAQ must be a list or contain an 'items'/'faq' list")


def _load_xlsx_rows(path: Path) -> List[Dict[str, Any]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl is required to load .xlsx FAQ files") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: List[Dict[str, Any]] = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        sheet_rows = list(worksheet.iter_rows(values_only=True))
        if not sheet_rows:
            continue
        headers = [_stringify(value) for value in sheet_rows[0]]
        for raw_row in sheet_rows[1:]:
            row = {
                headers[index]: value
                for index, value in enumerate(raw_row)
                if index < len(headers) and headers[index]
            }
            rows.append(row)
    return rows


def _normalize_rows(
    rows: Iterable[Dict[str, Any]], default_source: str
) -> List[Dict[str, str]]:
    normalized: List[Dict[str, str]] = []
    resolved_question_key: str | None = None
    resolved_answer_key: str | None = None

    raw_rows = list(rows)
    for row in raw_rows:
        question_key = _resolve_key(row, QUESTION_KEYS)
        answer_key = _resolve_key(row, ANSWER_KEYS)
        if question_key:
            resolved_question_key = question_key
        if answer_key:
            resolved_answer_key = answer_key

    missing = []
    if resolved_question_key is None:
        missing.append("Question")
    if resolved_answer_key is None:
        missing.append("Answer")
    if missing:
        raise ValueError(f"Missing required FAQ columns: {', '.join(missing)}")

    for row in raw_rows:
        question = _stringify(row.get(resolved_question_key))
        answer = _stringify(row.get(resolved_answer_key))
        if not question or not answer:
            continue

        category_key = _resolve_key(row, CATEGORY_KEYS)
        source_key = _resolve_key(row, SOURCE_KEYS)
        normalized.append(
            {
                "question": question,
                "answer": answer,
                "category": (
                    _stringify(row.get(category_key)) if category_key else "general"
                ),
                "source": (
                    _stringify(row.get(source_key)) if source_key else default_source
                ),
            }
        )

    return normalized


def _resolve_key(row: Dict[str, Any], aliases: set[str]) -> str | None:
    for key in row.keys():
        if _canonicalize(key) in aliases:
            return key
    return None


def _canonicalize(value: Any) -> str:
    return _stringify(value).strip().lower().replace("_", " ")


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
